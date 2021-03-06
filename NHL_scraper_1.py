import re
import os
import os.path
import openpyxl
import pandas
from urllib.request import urlopen
from bs4 import BeautifulSoup
from xlrd import open_workbook

def download_roster(stats_file_path, search_term, sheet_name):

    if os.path.exists(stats_file_path):
        ss = openpyxl.load_workbook(stats_file_path)
    else:
        ss = openpyxl.Workbook()

        # if there is a sheet with this name already, remove it
    try:
        sheet = ss.get_sheet_by_name(sheet_name)
        ss.remove_sheet(sheet)
    except:
        pass

        # create sheet for new data
    sheet = ss.create_sheet(sheet_name)

    html = urlopen('https://www.cbssports.com/nhl/stats/playersort/nhl/year-2017-season-regularseason-category-' \
                   + search_term + '?print_rows=9999').read()

    soup = BeautifulSoup(html, features = 'lxml')

    table = soup.find("table", {"class": "data"})

    for row in table.findAll('tr', {"class": "label"}):
        cells = row.findAll('th')
        cellTexts = []
        for col in cells:
            cellTexts.append(col.get_text())
        sheet.append(cellTexts)

    for row in table.findAll('tr', {"class": ["row1", "row2"]}):
        cells = row.findAll('td')
        cellTexts = []
        for col in cells:
            cellTexts.append(col.get_text())
        sheet.append(cellTexts)

    ss.save(stats_file_path)

def downloadStats(stats_file_path):
    download_roster(stats_file_path, 'points', 'points')
    download_roster(stats_file_path, 'wins', 'wins')
    download_roster(stats_file_path, 'penaltyminutes', 'penaltyminutes')

def display_menu():
    print("==================== MENU ====================")
    print("1. Display Points Leaders")
    print("2. Display Wins Leaders")
    print("3. Display Penalty Minutes Leaders")
    print("r. Refresh data")
    print("q. Quit")
    option = input("Please select an option: ")
    return option

def getCountChoice():
    done = False
    while (done == False):
        countStr = input("Enter the maximum number of players to display (0 for all): ")
        try:
            top = int(countStr)
            done = True;
        except:
            pass

    return top

def printRow(row, index):
    if index == 0:
        print('Rank, ', end='')
    else:
        print(str(index) + ". ", end='')
    values = []
    for v in row:
        values.append(v.value)
    print(", ".join(values))
    # for value in row:
    #     print(value.value + ", ", end='')
    #print()

def printNRows(rows, n):
    if n == 0:
        rcount = 0
        for row in rows:
            printRow(row, rcount)
            rcount += 1
        return

    rcount = 0
    for row in rows:
        printRow(row, rcount)
        if rcount == n:
            return
        rcount += 1


#print("cwd: ", os.getcwd())
#print("look for excel file at", os.getcwd())


def display_stats(stats_file_path, sheet_name, top = 10):
    print("********************", sheet_name.upper(), "********************")
    workbook = open_workbook(stats_file_path)
    sheet = workbook.sheet_by_name(sheet_name)
    rows = sheet.get_rows()
    printNRows(rows, top)


def main():
    dir = 'downloads/stats'
    if not os.path.exists(dir):
        os.makedirs(dir)

    stats_file_path = dir + "/nhl_stats.xlsx"
    downloadStats(stats_file_path)

    while True:
        top = 0
        option = display_menu()
        if (option.lower() == 'q'):
            break
        elif option in ['1', '2', '3']:
            if option == '1':
                option = 'points'
            if option == '2':
                option = 'wins'
            if option == '3':
                option = 'penaltyminutes'
            top = getCountChoice()
            display_stats(stats_file_path, option, top)
        elif (option.lower() == 'r'):
            downloadStats(stats_file_path)
        else:
            pass

if __name__ == '__main__':
    main()
